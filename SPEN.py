import os
import re
import difflib
import io
import calendar as py_calendar
from datetime import datetime
 
import pandas as pd
import streamlit as st
import plotly.express as px
import plotly.graph_objects as go
from streamlit_calendar import calendar as st_calendar
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
 
st.set_page_config(page_title="Network Job Tracker", layout="wide")
 
# ============================================================
# OUTAGES PROGRAMME (uploaded by the user - not read from the network)
# ============================================================
# Streamlit Cloud has no access to internal UNC/network paths, so the
# workbook is uploaded via a file_uploader instead of read from disk.
 
 
@st.cache_data(show_spinner=False, max_entries=5)
def build_calendar_events(cal_df: pd.DataFrame) -> list:
    """Vectorized FullCalendar event build (zip over plain Python lists,
    no .iterrows()) - cached so re-rendering the calendar on an unrelated
    rerun (a date click now costs a full script rerun) doesn't rebuild
    the same event list from scratch every time."""
    palette = ["#2563eb", "#dc2626", "#059669", "#d97706", "#7c3aed", "#0891b2", "#be185d", "#4d7c0f"]
    districts_sorted = sorted(cal_df["District"].dropna().unique())
    district_colors = {d: palette[i % len(palette)] for i, d in enumerate(districts_sorted)}
 
    starts = cal_df["Outage Date"].dt.strftime("%Y-%m-%d").tolist()
    districts = cal_df["District"].tolist()
    schemes = cal_df["Scheme"].tolist()
    outage_nums = cal_df["Outage #"].tolist()
    circuits_evt = cal_df["Circuit"].tolist()
    pids_evt = cal_df["PID"].tolist()
    pms_evt = cal_df["SPEN PM"].tolist()
    pois_evt = cal_df["POI"].tolist()
 
    def _evt_str(v):
        return "" if pd.isna(v) or str(v).strip() == "" else str(v).strip()
 
    events = []
    for start, district, scheme, onum, circuit, pid, pm, poi in zip(
        starts, districts, schemes, outage_nums, circuits_evt, pids_evt, pms_evt, pois_evt
    ):
        d_str, s_str = _evt_str(district), _evt_str(scheme)
        # Full title (District — Scheme — PID — Circuit — Outage # — PM — POI): this is
        # what FullCalendar puts on the event element's native "title" attribute, so
        # hovering shows the complete text even though the day cell itself clips it to
        # one line (see custom_css at the call site) instead of wrapping and breaking
        # the grid.
        parts = [d_str, s_str]
        for label, val in [("PID", pid), ("Circuit", circuit), ("Outage #", onum), ("PM", pm), ("POI", poi)]:
            v = _evt_str(val)
            if v:
                parts.append(f"{label} {v}")
        full_title = " — ".join(p for p in parts if p) or "Outage"
        color = district_colors.get(district, "#2563eb")
        events.append({
            "title": full_title,
            "start": start,
            "allDay": True,
            "backgroundColor": color,
            "borderColor": color,
        })
    return events
 
 
def _outage_full_title(row) -> str:
    """Same 'District — Scheme — PID — Circuit — Outage # — PM — POI' text the
    on-screen calendar puts in each event's title (see build_calendar_events),
    but built from a single row instead of parallel lists - used for the
    downloadable export where every name needs to be fully spelled out
    on the page, not just available on hover."""
    def s(v):
        return "" if pd.isna(v) or str(v).strip() == "" else str(v).strip()
 
    parts = [s(row.get("District")), s(row.get("Scheme"))]
    for label, key in [("PID", "PID"), ("Circuit", "Circuit"), ("Outage #", "Outage #"), ("PM", "SPEN PM"), ("POI", "POI")]:
        v = s(row.get(key))
        if v:
            parts.append(f"{label} {v}")
    return " — ".join(p for p in parts if p) or "Outage"
 
 
@st.cache_data(show_spinner="Building calendar workbook...", max_entries=3)
def build_full_calendar_excel(cal_df: pd.DataFrame) -> bytes:
    """Builds a downloadable .xlsx with the outage programme in the same
    visual layout as the on-screen calendar (a month grid, Mon-Sun columns,
    one box per day) - but every event's full name is written out in the
    cell instead of being clipped to one line with the rest only on hover.
    Also includes a flat 'All outages' list sheet as a plain-text safety net.
 
    Returns raw workbook bytes, cached on the filtered dataframe's content
    so re-downloading after an unrelated widget interaction doesn't rebuild
    the whole file from scratch."""
    cal_df = cal_df.dropna(subset=["Outage Date"]).copy()
 
    header_font = Font(bold=True, color="FFFFFF", size=13)
    header_fill = PatternFill("solid", fgColor="1E3A8A")
    weekday_fill = PatternFill("solid", fgColor="334155")
    weekday_font = Font(bold=True, color="FFFFFF")
    outside_fill = PatternFill("solid", fgColor="F1F5F9")
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    day_num_font = Font(bold=True, size=10, color="1E293B")
    event_alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
 
    wb = Workbook()
    wb.remove(wb.active)
 
    # ---- Flat list sheet first (every row, every column spelled out) ----
    ws_list = wb.create_sheet("All outages (list)")
    list_cols = ["Outage Date", "Weekday", "District", "Scheme", "Outage #", "Circuit", "PID", "SPEN PM", "POI", "Full name"]
    for j, colname in enumerate(list_cols, start=1):
        c = ws_list.cell(row=1, column=j, value=colname)
        c.font = weekday_font
        c.fill = weekday_fill
    list_df = cal_df.sort_values("Outage Date").copy()
    list_df["Full name"] = list_df.apply(_outage_full_title, axis=1)
    for i, row in enumerate(list_df.itertuples(index=False), start=2):
        row_map = row._asdict() if hasattr(row, "_asdict") else dict(zip(list_df.columns, row))
        for j, colname in enumerate(list_cols, start=1):
            val = row_map.get(colname, "")
            if hasattr(val, "strftime"):
                val = val.strftime("%Y-%m-%d")
            ws_list.cell(row=i, column=j, value=val)
    widths = {"Outage Date": 14, "Weekday": 12, "District": 16, "Scheme": 22, "Outage #": 12,
              "Circuit": 16, "PID": 12, "SPEN PM": 18, "POI": 16, "Full name": 70}
    for j, colname in enumerate(list_cols, start=1):
        ws_list.column_dimensions[get_column_letter(j)].width = widths.get(colname, 16)
    ws_list.freeze_panes = "A2"
 
    if cal_df.empty:
        bio = io.BytesIO()
        wb.save(bio)
        return bio.getvalue()
 
    cal_df["_year"] = cal_df["Outage Date"].dt.year
    cal_df["_month"] = cal_df["Outage Date"].dt.month
    cal_df["_full_title"] = cal_df.apply(_outage_full_title, axis=1)
 
    months_present = sorted(cal_df[["_year", "_month"]].drop_duplicates().itertuples(index=False, name=None))
    weekday_labels = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
 
    for year, month in months_present:
        sheet_name = f"{py_calendar.month_abbr[month]} {year}"[:31]
        ws = wb.create_sheet(sheet_name)
 
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
        title_cell = ws.cell(row=1, column=1, value=f"{py_calendar.month_name[month]} {year}")
        title_cell.font = header_font
        title_cell.fill = header_fill
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 24
 
        for i, wd in enumerate(weekday_labels, start=1):
            c = ws.cell(row=2, column=i, value=wd)
            c.font = weekday_font
            c.fill = weekday_fill
            c.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(i)].width = 32
 
        month_events = cal_df[(cal_df["_year"] == year) & (cal_df["_month"] == month)]
        events_by_day = month_events.groupby(month_events["Outage Date"].dt.day)["_full_title"].apply(list).to_dict()
 
        weeks = py_calendar.Calendar(firstweekday=0).monthdatescalendar(year, month)
        row_idx = 3
        for week in weeks:
            max_lines = 1
            for col_idx, day_date in enumerate(week, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = border
                cell.alignment = event_alignment
                if day_date.month == month:
                    day_events = events_by_day.get(day_date.day, [])
                    text = str(day_date.day)
                    if day_events:
                        text += "\n" + "\n".join(f"• {e}" for e in day_events)
                    cell.value = text
                    max_lines = max(max_lines, 1 + len(day_events))
                else:
                    cell.value = ""
                    cell.fill = outside_fill
            ws.row_dimensions[row_idx].height = max(22, 15 * max_lines)
            row_idx += 1
 
        ws.freeze_panes = "A3"
 
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()
 
 
@st.cache_data(show_spinner="Reading outage programme...", max_entries=3)
def load_outage_programme(file_bytes: bytes):
    """Cached on the uploaded file's bytes - re-parses only when a
    different file (or a changed version of the same file) is uploaded,
    not on every rerun/widget interaction.
 
    Returns (df, error_message). Never raises - if the uploaded workbook
    doesn't have a '2026' sheet in the expected layout, this should show a
    clean st.error rather than crash the whole app."""
    try:
        df = pd.read_excel(
            io.BytesIO(file_bytes),
            sheet_name="2026",
            header=6,                     # Excel row 7 is the header row (0-indexed = 6)
            usecols="A,B,C,E,F,G,L,M,N",  # District, Outage Date, Weekday, Scheme, Outage #, Circuit, PID, SPEN PM, POI
            engine="openpyxl",
        )
        df.columns = [
            "District", "Outage Date", "Weekday", "Scheme",
            "Outage #", "Circuit", "PID", "SPEN PM", "POI",
        ]
        df = df.dropna(how="all")
        df["Outage Date"] = pd.to_datetime(df["Outage Date"], errors="coerce")
        return df, None
    except Exception as e:
        return None, str(e)
 
 
# ============================================================
# YOUR MAPPING DICTIONARIES
# ============================================================
CV7_erect = {
    "Erect Single HV/EHV Pole, up to and including 12 metre pole": "CV7 HV pole",
    "Erect Single HV/EHV Pole, up to and including 12 metre pole.": "CV7 HV pole",
}
CV7_erect_H = {
    "Erect Section Structure 'H' HV/EHV Pole, up to and including 12 metre pole.": "CV7 HV pole"
}
CV7_erect_lv = {
    "Erect LV Structure Single Pole, up to and including 12 metre pole": "CV7 LV pole",
}
CV7_recover = {
    "Recover single pole, up to and including 15 metres in height, and reinstate, all ground conditions": "CV7",
    "Recover 'A' / 'H' pole, up to and including 15 metres in height, and reinstate, all ground conditions": "CV7 HV pole",
}
CV7_Tx = {
    "Erect pole mounted transformer up to 100kVA 1.ph.": "CV7 Tx",
    "Erect pole mounted transformer up to 200kVA 3.p.h.": "CV7 Tx",
    "Erect Voltage Regulator.": "CV7 Tx",
    "Erect Voltage Transformer (VT), RTU or Repeater": "CV7 Tx",
    "Erect 12kV/36kV Surge arrestors ( directly mounted ).": "CV7 Tx",
    "Remove pole mounted tranformer.": "CV7 Tx",
    "Remove platform mounted or 'H' pole mounted transformer.": "CV7 Tx",
}
transformer = {
    "Transformer 1ph 50kVA": "TX 1ph (50kVA)",
    "Transformer 3ph 50kVA": "TX 3ph (50kVA)",
    "Transformer 1ph 100kVA": "TX 1ph (100kVA)",
    "Transformer 1ph 25kVA": "TX 1ph (25kVA)",
    "Transformer 3ph 200kVA": "TX 3ph (200kVA)",
    "Transformer 3ph 100kVA": "TX 3ph (100kVA)",
}
CV7_OHL_CONDUCTOR_instal = {
    "Install bare conductor, run out, sag, terminate, bind in and connect jumpers; <100mm²": "CV7 OHL CONDUCTOR",
    "Install bare conductor, run out, sag, terminate, bind in and connect jumpers; >=100mm² <200mm²": "CV7 OHL CONDUCTOR",
    "Install conductor, run out, sag, terminate, clamp in and form jumper loops; >=200mm²": "CV7 OHL CONDUCTOR",
}
CV7_OHL_CONDUCTOR_recover = {
    "Recover overhead wire and fittings; HV/EHV overhead line or Hardex Pilot (1 conductor)": "CV7 OHL CONDUCTOR",
    "Recover overhead wire and fittings; HV/EHV overhead line or Hardex Pilot (2 conductor)": "CV7 OHL CONDUCTOR",
    "Recover overhead wire and fittings; HV/EHV overhead line or Hardex Pilot (3 conductor)": "CV7 OHL CONDUCTOR",
}
CV7_OHL_CONDUCTOR_LV_instal = {
    "Install conductor, run out, sag, terminate, clamp in and connect jumpers; 2c": "CV7 OHL CONDUCTOR LV",
    "Install conductor, run out, sag, terminate, clamp in and connect jumpers; 4c": "CV7 OHL CONDUCTOR LV",
    "Install conductor, run out, sag, terminate, clamp in and connect jumpers; 2c + Earth": "CV7 OHL CONDUCTOR LV",
    "Install conductor, run out, sag, terminate, clamp in and connect jumpers; 4c + Earth": "CV7 OHL CONDUCTOR LV",
}
CV7_OHL_CONDUCTOR_LV_recover = {
    "Recover overhead wires and fittings; LV openwire overhead line (2 conductors)": "CV7 OHL CONDUCTOR LV",
    "Recover overhead wires and fittings; LV openwire overhead line (3 conductors)": "CV7 OHL CONDUCTOR LV",
    "Recover overhead wires and fittings; LV openwire overhead line (4 conductors)": "CV7 OHL CONDUCTOR LV",
    "Recover overhead wires and fittings; LV openwire overhead line (5 conductors)": "CV7 OHL CONDUCTOR LV",
    "Recover overhead wires and fittings; LV service overhead line (open, concentric or ABC, 2 conductors)": "CV7 OHL CONDUCTOR LV",
    "Recover overhead wires and fittings; LV service overhead line (open, concentric or ABC, 3 conductors)": "CV7 OHL CONDUCTOR LV",
    "Recover overhead wires and fittings; LV service overhead line (open, concentric or ABC, 4 conductors)": "CV7 OHL CONDUCTOR LV",
    "Recover overhead wires and fittings; LV service overhead line (open, concentric or ABC, 5 conductors)": "CV7 OHL CONDUCTOR LV",
    "Recover cleated service": "CV7 OHL CONDUCTOR LV",
}
Switch = {
    "Noja": "Noja",
    "11kV PMSW (Soule)": "11kV PMSW (Soule)",
    "11kv ABSW Hookstick Standard": "11kv ABSW Hookstick Standard",
    "11kv ABSW Hookstick Spring loaded mech": "11kv ABSW Hookstick Spring loaded mech",
    "33kv ABSW Hookstick Dependant": "33kv ABSW Hookstick Dependant",
}
Fuses = {
    "100A LV Fuse JPU 82.5mm": "100A LV Fuse JPU 82.5mm",
    "160A LV Fuse JPU 82.5mm": "160A LV Fuse JPU 82.5mm",
    "200A LV Fuse JPU 82.5mm": "200A LV Fuse JPU 82.5mm",
    "315A LV Fuse JPU 82.5mm": "315A LV Fuse JPU 82.5mm",
    "400A LV Fuse JPU 82.5mm": "400A LV Fuse JPU 82.5mm",
    "200A LV Fuse JSU 92mm": "200A LV Fuse JSU 92mm",
    "315A LV Fuse JSU 92mm": "315A LV Fuse JSU 92mm",
    "400A LV Fuse JSU 92mm": "400A LV Fuse JSU 92mm",
    "100A LV Fuse - Porcelain screw-in": "100A LV Fuse - Porcelain screw-in",
    "160A LV Fuse - Porcelain screw-in": "160A LV Fuse - Porcelain screw-in",
    "200A LV Fuse - Porcelain screw-in": "200A LV Fuse - Porcelain screw-in",
    "Single Phase cut out kit 100A Henley Series 7": "Single Phase cut out kit 100A Henley Series 7",
    "Three Phase cut out kit 100A Henley Series 7": "Three Phase cut out kit 100A Henley Series 7",
    "Three Phase 200A Cut out": "Three Phase 200A Cut out",
    "Cut out Fuse (MF) 60A": "Cut out Fuse (MF) 60A",
    "Cut out Fuse (MF) 80A": "Cut out Fuse (MF) 80A",
    "Cut out Fuse (MF) 100A": "Cut out Fuse (MF) 100A",
    "11KV FUSE UNIT - C-TYPE": "11KV FUSE UNIT - C-TYPE",
    "11KV SOLID LINK - C-TYPE": "11KV SOLID LINK - C-TYPE",
    "11KV OHL ASL C-TYPE RESET 20A 2 SHOT": "11KV OHL ASL C-TYPE RESET 20A 2 SHOT",
    "11KV OHL ASL C-TYPE RESET 25A 2 SHOT": "11KV OHL ASL C-TYPE RESET 25A 2 SHOT",
    "11KV OHL ASL C-TYPE RESET 40A 1 SHOT": "11KV OHL ASL C-TYPE RESET 40A 1 SHOT",
    "11KV OHL ASL C-TYPE RESET 40A 2 SHOT": "11KV OHL ASL C-TYPE RESET 40A 2 SHOT",
    "11KV OHL ASL C-TYPE RESET 63A 1 SHOT": "11KV OHL ASL C-TYPE RESET 63A 1 SHOT",
    "11KV OHL ASL C-TYPE RESET 63A 2 SHOT": "11KV OHL ASL C-TYPE RESET 63A 2 SHOT",
    "11KV OHL ASL C-TYPE RESET 63A 3 SHOT": "11KV OHL ASL C-TYPE RESET 63A 3 SHOT",
    "11KV OHL ASL C-TYPE RESET 100A 1 SHOT": "11KV OHL ASL C-TYPE RESET 100A 1 SHOT",
    "11KV OHL ASL C-TYPE RESET 100A 2 SHOT": "11KV OHL ASL C-TYPE RESET 100A 2 SHOT",
    "11KV OHL ASL C-TYPE RESET 100A 3 SHOT": "11KV OHL ASL C-TYPE RESET 100A 3 SHOT",
    "11KV OHL FUSE ELEMENT C-TYPE 15A": "11KV OHL FUSE ELEMENT C-TYPE 15A",
    "11KV OHL FUSE ELEMENT C-TYPE 25A": "11KV OHL FUSE ELEMENT C-TYPE 25A",
    "11KV OHL FUSE ELEMENT C-TYPE 30A": "11KV OHL FUSE ELEMENT C-TYPE 30A",
    "11KV OHL FUSE ELEMENT C-TYPE 40A": "11KV OHL FUSE ELEMENT C-TYPE 40A",
    "11KV OHL FUSE ELEMENT C-TYPE 50A": "11KV OHL FUSE ELEMENT C-TYPE 50A",
    "11KV OHL ASL DJP-TYPE 20A 2 SHOT": "11KV OHL ASL DJP-TYPE 20A 2 SHOT",
    "11KV OHL ASL DJP-TYPE 25A 1 SHOT": "11KV OHL ASL DJP-TYPE 25A 1 SHOT",
    "11KV OHL ASL DJP-TYPE 25A 2 SHOT": "11KV OHL ASL DJP-TYPE 25A 2 SHOT",
    "11KV OHL ASL DJP-TYPE 40A 1 SHOT": "11KV OHL ASL DJP-TYPE 40A 1 SHOT",
    "11KV OHL ASL DJP-TYPE 40A 2 SHOT": "11KV OHL ASL DJP-TYPE 40A 2 SHOT",
    "11KV OHL ASL DJP-TYPE 63A 1 SHOT": "11KV OHL ASL DJP-TYPE 63A 1 SHOT",
    "11KV OHL ASL DJP-TYPE 63A 2 SHOT": "11KV OHL ASL DJP-TYPE 63A 2 SHOT",
    "11KV OHL ASL DJP-TYPE 63A 3 SHOT": "11KV OHL ASL DJP-TYPE 63A 3 SHOT",
    "11KV OHL ASL DJP-TYPE 100A 1 SHOT": "11KV OHL ASL DJP-TYPE 100A 1 SHOT",
    "11KV OHL ASL DJP-TYPE 100A 2 SHOT": "11KV OHL ASL DJP-TYPE 100A 2 SHOT",
    "11KV OHL ASL DJP-TYPE 100A 3 SHOT": "11KV OHL ASL DJP-TYPE 100A 3 SHOT",
    "11KV OHL FUSE ELEMENT DJP-TYPE 15A": "11KV OHL FUSE ELEMENT DJP-TYPE 15A",
    "11KV OHL FUSE ELEMENT DJP-TYPE 25A": "11KV OHL FUSE ELEMENT DJP-TYPE 25A",
    "11KV OHL FUSE ELEMENT DJP-TYPE 30A": "11KV OHL FUSE ELEMENT DJP-TYPE 30A",
    "11KV OHL FUSE ELEMENT DJP-TYPE 40A": "11KV OHL FUSE ELEMENT DJP-TYPE 40A",
    "11KV OHL FUSE ELEMENT DJP-TYPE 50A": "11KV OHL FUSE ELEMENT DJP-TYPE 50A",
}
CV31 = {
    "Replace / Fit safety or warning sign, number plates or name plate": "CV31",
    "Barbed Wire Wrap ACD (or Enhanced) single pole or stay - Replace/Repair": "CV31",
    "Steelwork bonding repair / fit.": "CV31",
    "Replace LV/HV/Earth guard missing / damaged.": "CV31",
}
CV8 = {
    "Tighten existing stay.": "CV8",
    "Erect/Replace stay above ground only.": "CV8",
    "Erect/Replace stay complete including block or driven type anchor": "CV8",
    "Erect/Replace stay complete including rock type anchor": "CV8",
    "Retrofit structure with Anchor Clamp fitting for Section / Angle / Terminal support": "CV8",
    "Erect Single Crossarm to single pole.": "CV8",
    "Erect Double Crossarm 'H' Pole formation": "CV8",
    "Remove Steelwork crossarm item only": "CV8",
    "Change 11kV Insulators to avoid contamination from old conductor": "CV8",
    "Change 33kV Insulators to avoid contamination from old conductor": "CV8",
    "Replace tension insulator, 11kV.": "CV8",
    "Replace tension insulator, 33kV.": "CV8",
    "Additional cost for fitting Stay Outrigger Bracket": "CV8",
    "Additional cost for fitting Angle / Terminal stay attachment plates on Heavy Construction as SP4009862": "CV8",
    "Recover and reinstate stay position,all ground conditions.": "CV8",
    "Fit foundation block to existing pole.": "CV8",
    "Fit bog shoe foundation to existing single pole.": "CV8",
    "Replace jumper / dropper mechanical connection with compression connection": "CV8",
    "Replace jumper / dropper with live line bail and flexible jumper conductor": "CV8",
    "Replace / Repair conductor with mid span joint using compression connection": "CV8",
    "Conductor repair; piece in conductor including compression joints": "CV8",
    "Bind In Conductors; 1.ph 11kV Intermediate / Pin Angle pole.": "CV8",
    "Bind In Conductors; 3.ph 11kV Intermediate / Pin Angle pole.": "CV8",
    "Conductor Terminations - 1.ph 11kV Section pole including jumpers.": "CV8",
    "Conductor Terminations - 3.ph 11kV Section pole including jumpers.": "CV8",
    "Conductor Terminations - 1.ph 11kV Terminal pole.": "CV8",
    "Conductor Terminations - 3.ph 11kV Terminal pole.": "CV8",
    "Unbind and reregulate existing conductors": "CV8",
    "Convert 1.ph 11kV Intermediate pole into Section Pole.": "CV8",
    "Convert 1.ph/3.p.h. 11kV line pole into Terminal Pole.": "CV8",
    "Convert 3.ph 11kV Intermediate pole into Section Pole.": "CV8",
    "Replace 11kV/33kV insulator pin and insulator, including unbinding and binding in": "CV8",
    "Replace 11kV/33kV insulator binder": "CV8",
    "Replace tension insulator, 11kV": "CV8",
    "Replace tension insulator, 33kV": "CV8",
    "Replace 11kV/33kV dead end termination": "CV8",
    "Additional cost for erection of pilot pin and insulator or pilot post insulator (11kV or 33kV)": "CV8",
    "Replace insulated conductor HV/LV earth above ground to first rod": "CV8",
    "Install Copper Covered Green / Yellow HV Earth or Black LV Earth to foot of pole": "CV8",
    "Install EHV/ HV Earth Electrode including excavate & reinstate (up to 8mtrs)": "CV8",
    "Install LV Earth Electrode including excavate & reinstate (up to 28mtrs)": "CV8",
    "Additional extra over for additional earthing excavated, laid & backfilled": "CV8",
    "Install Earth Electrode within cable trench": "CV8",
    "Erect 11kV Cable Termination ( incorporating surge arrestors )": "CV8",
    "Erect 33kV Cable Termination ( incorporating surge arrestors )": "CV8",
    "Steelwork bonding repair / fit": "CV8",
    "Erect 1.ph LV cable pole termination": "CV8",
    "Erect 3.ph LV cable pole termination": "CV8",
    "Remove 11kV/33kV Cable termination": "CV8",
    "Remove LV cable termination": "CV8",
    "Repair pole twist - including unbind / rebind.": "CV8",
}
 
POLE_CATEGORIES = {
    "CV7_erect": CV7_erect,
    "CV7_erect_H": CV7_erect_H,
    "CV7_erect_lv": CV7_erect_lv,
    "CV7_recover": CV7_recover,
}
 
# Friendly labels for the pole chart/cards, in the "Display (technical_name)"
# style so it's obvious which export-tool category each bar corresponds to.
POLE_DISPLAY_NAMES = {
    "CV7_erect": "CV7 (CV7_erect)",
    "CV7_erect_H": "CV7 H Pole (CV7_erect_H)",
    "CV7_erect_lv": "CV7 LV Pole (CV7_erect_lv)",
    "CV7_recover": "CV7 Recover (CV7_recover)",
}
 
# NOTE: CV7_SWITCHGEAR / CV7_UG / CV7_CB were removed from ALL_CATEGORIES.
# In the export tool, `categories` gets redefined a second time and that
# second definition (the one actually used to build sheets/Summary) never
# includes these three - so the exporter never produces a "true" value for
# them. Showing cards for them here would just be comparing against
# nothing. If you do want them tracked, they need to be added back into
# the export tool's `categories`/`extra_categories` lists first.
ALL_CATEGORIES = {
    **POLE_CATEGORIES,
    "CV7_Tx": CV7_Tx,
    "transformer": transformer,
    "CV7_OHL_CONDUCTOR_instal": CV7_OHL_CONDUCTOR_instal,
    "CV7_OHL_CONDUCTOR_recover": CV7_OHL_CONDUCTOR_recover,
    "CV7_OHL_CONDUCTOR_LV_instal": CV7_OHL_CONDUCTOR_LV_instal,
    "CV7_OHL_CONDUCTOR_LV_recover": CV7_OHL_CONDUCTOR_LV_recover,
    "Switch": Switch,
    "Fuses": Fuses,
    "CV31": CV31,
    "CV8": CV8,
}
 
# Categories that use the exporter's process_cv() logic instead of a plain
# sum: dedupe by pole, drop zero-qty rows, exclude poles already counted
# under a CV7 erect/recover item, then COUNT distinct poles (not sum qty).
POLE_DEDUPE_CATEGORIES = {"CV8", "CV31"}
 
# Switch is split into three sub-types for display instead of one lump
# card. Each key is the display name, each value is the list of raw
# descriptions (as written in the Switch dict above) that count toward it.
SWITCH_SUBTYPES = {
    "Noja": ["Noja"],
    "Soule": ["11kV PMSW (Soule)"],
    "ABSW": [
        "11kv ABSW Hookstick Standard",
        "11kv ABSW Hookstick Spring loaded mech",
        "33kv ABSW Hookstick Dependant",
    ],
}
 
# ============================================================
# IMAGE GROUPS for the Mapped Items tab
# Images live in an "Images" folder alongside this script (same repo path).
# "image": None means the group is shown without an image (no warning).
# Any category not listed in a group's "categories" (or covered by
# "subtypes") falls through to the "Other items" section at the end.
# ============================================================
IMAGE_DIR = "Images"
CARD_GROUPS = [
    {
        "title": "Poles",
        "image": os.path.join(IMAGE_DIR, "Poles.png"),
        "categories": ["CV7_recover", "CV7_erect", "CV7_erect_H", "CV7_erect_lv"],
    },
    {
        "title": "Transformers",
        "image": os.path.join(IMAGE_DIR, "Transformer.png"),
        "categories": ["CV7_Tx", "transformer"],
    },
    {
        "title": "Conductor",
        "image": os.path.join(IMAGE_DIR, "Cable.png"),
        "categories": [
            "CV7_OHL_CONDUCTOR_instal",
            "CV7_OHL_CONDUCTOR_recover",
            "CV7_OHL_CONDUCTOR_LV_instal",
            "CV7_OHL_CONDUCTOR_LV_recover",
        ],
    },
    {
        "title": "Switch gear",
        "image": os.path.join(IMAGE_DIR, "Switchgear.png"),
        "subtypes": SWITCH_SUBTYPES,
    },
]
 
HV_POLE_KEY = "Recover 'A' / 'H' pole, up to and including 15 metres in height, and reinstate, all ground conditions"
HV_POLE_MULTIPLIER = 2
 
# ============================================================
# HELPERS (aligned with the export script's normalization)
# ============================================================
def normalize_item(x):
    if pd.isna(x):
        return ""
    s = str(x).replace("\u200b", "").replace("\u200e", "").replace("\u200f", "").replace("\xa0", "").strip().upper()
    return re.sub(r"\s+", " ", s)
 
 
def normalize_pole(p):
    """Mirrors the export tool's normalize_pole(): strips zero-width/
    directional/nbsp characters, uppercases, and removes ALL whitespace
    (not just collapsing it) so pole IDs compare exactly like the exporter."""
    if pd.isna(p):
        return ""
    s = str(p).replace("\u200b", "").replace("\u200e", "").replace("\u200f", "").replace("\xa0", "").strip().upper()
    return re.sub(r"\s+", "", s)
 
 
def clean_job(value):
    """Strip leading 'C -'/'M -' prefix, cut at 'map', drop SPxxxx/GSPxxxx/bare digits."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    text = str(value).strip()
    text = re.sub(r'^[A-Za-z]\s*-\s*', '', text)
    m = re.search(r'map', text, flags=re.IGNORECASE)
    if m:
        text = text[: m.start()]
    text = re.sub(r'\bGSP\d+\b', '', text, flags=re.IGNORECASE)
    text = re.sub(r'\bSP\d+\b', '', text, flags=re.IGNORECASE)
    text = re.sub(r'\b\d+\b', '', text)
    text = re.sub(r'\s{2,}', ' ', text)
    text = re.sub(r'^[\s\-\u2013_,.:]+|[\s\-\u2013_,.:]+$', '', text)
    return text.strip()
 
 
# Native unit each category's qsub is recorded in - used to format the
# mapped-item cards for conductor lengths.
UNIT_CONFIG = {
    "CV7_OHL_CONDUCTOR_recover": "m",
    "CV7_OHL_CONDUCTOR_LV_recover": "m",
    "CV7_OHL_CONDUCTOR_instal": "km",
    "CV7_OHL_CONDUCTOR_LV_instal": "km",
}
 
 
def format_length(value, native_unit):
    """Converts value (in native_unit) to meters, then displays in km if
    the meters equivalent is >=1000, otherwise in meters."""
    meters = value * 1000 if native_unit == "km" else value
    if meters >= 1000:
        return f"{meters / 1000:,.2f} km"
    return f"{meters:,.0f} m"
 
 
def dedupe_jobs(values, threshold=0.65):
    kept, is_dup = [], []
    for v in values:
        if not v:
            is_dup.append(False)
            continue
        hit = any(difflib.SequenceMatcher(None, v.lower(), k.lower()).ratio() >= threshold for k in kept)
        is_dup.append(hit)
        if not hit:
            kept.append(v)
    return is_dup
 
 
def show_total_banner(label, value_str):
    """A large, centered KPI number - meant to sit directly under a chart's
    subheader so the total is the first thing seen after the title."""
    st.markdown(
        f"""
        <div style="text-align:center; padding: 4px 0 18px 0;">
            <div style="font-size:2.6rem; font-weight:800; color:#1e3a8a; line-height:1.15;">{value_str}</div>
            <div style="font-size:0.9rem; color:#475569; text-transform:uppercase; letter-spacing:0.05em;">{label}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )
 
 
@st.cache_data(show_spinner=False, max_entries=8)
def build_pole_task_table(day_df: pd.DataFrame, item_col: str, qsub_col: str, pole_col) -> pd.DataFrame:
    """Work Instructions for a single date, as a table instead of a Word
    doc - same grouping the workpacks tool's generate_pole_docx() uses
    (one row per pole+task, pole-first order, blank/'0' rows dropped,
    'Erect' flagged) but built with vectorized pandas ops instead of
    row-by-row iterrows(), and cached so re-selecting the same date on a
    later rerun (e.g. from an unrelated widget elsewhere) is instant.
    """
    if day_df.empty or item_col not in day_df.columns or qsub_col not in day_df.columns:
        return pd.DataFrame(columns=["Pole", "Task", "Qty", "Erect"])
 
    pole_vals = (
        day_df[pole_col].astype(str).str.strip()
        if pole_col and pole_col in day_df.columns
        else pd.Series("(no pole column mapped)", index=day_df.index)
    )
    out = pd.DataFrame({
        "Pole": pole_vals,
        "Task": day_df[item_col].astype(str).str.strip(),
        "Qty": day_df[qsub_col],
    })
 
    bad = {"", "nan", "none", "nat", "0"}
    out = out[~out["Pole"].str.lower().isin(bad) & ~out["Task"].str.lower().isin(bad)]
    if out.empty:
        return pd.DataFrame(columns=["Pole", "Task", "Qty", "Erect"])
 
    out["Erect"] = out["Task"].str.contains("erect", case=False, na=False).map({True: "Yes", False: ""})
 
    # pole-first order (same idea as the workpacks tool's _dedup_keep_order):
    # each pole's rows stay together, ordered by that pole's first appearance.
    pole_order = out["Pole"].drop_duplicates().tolist()
    out["Pole"] = pd.Categorical(out["Pole"], categories=pole_order, ordered=True)
    out = out.sort_values("Pole", kind="stable").reset_index(drop=True)
    out["Pole"] = out["Pole"].astype(str)
    return out
 
 
@st.cache_data(show_spinner=False, max_entries=8)
def render_pole_task_table_html(df: pd.DataFrame) -> str:
    """Renders build_pole_task_table()'s output with the Pole column
    merged (rowspan) across each pole's task rows and alternating
    shading per pole group - the HTML/table equivalent of the workpacks
    tool's Excel merge_column_runs()/pole banding, since st.dataframe has
    no way to merge cells. Erect tasks are shown in red/bold, same as
    the workpacks tool's docx highlighting."""
    if df.empty:
        return ""
 
    def esc(v):
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return ""
        return str(v).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
 
    poles = df["Pole"].tolist()
    tasks = df["Task"].tolist()
    qtys = df["Qty"].tolist()
    erects = df["Erect"].tolist()
    n = len(poles)
 
    band_colors = ["#ffffff", "#eef2f7"]
    rows_html = []
    i, band = 0, 0
    while i < n:
        j = i
        while j + 1 < n and poles[j + 1] == poles[i]:
            j += 1
        span = j - i + 1
        bg = band_colors[band % 2]
        band += 1
        for k in range(i, j + 1):
            erect_style = "color:#dc2626;font-weight:600;" if erects[k] == "Yes" else ""
            cells = []
            if k == i:
                cells.append(
                    f"<td rowspan='{span}' style='background:{bg};font-weight:600;"
                    f"vertical-align:top;padding:6px 10px;border:1px solid #e3e7ee;"
                    f"white-space:nowrap;'>{esc(poles[k])}</td>"
                )
            cells.append(
                f"<td style='background:{bg};padding:6px 10px;border:1px solid #e3e7ee;{erect_style}'>"
                f"{esc(tasks[k])}</td>"
            )
            cells.append(
                f"<td style='background:{bg};padding:6px 10px;border:1px solid #e3e7ee;"
                f"text-align:right;white-space:nowrap;'>{esc(qtys[k])}</td>"
            )
            cells.append(
                f"<td style='background:{bg};padding:6px 10px;border:1px solid #e3e7ee;"
                f"text-align:center;'>{esc(erects[k])}</td>"
            )
            rows_html.append("<tr>" + "".join(cells) + "</tr>")
        i = j + 1
 
    header = (
        "<tr style='background:#1e3a8a;color:#ffffff;'>"
        "<th style='padding:8px 10px;text-align:left;'>Pole</th>"
        "<th style='padding:8px 10px;text-align:left;'>Task (MD Poling)</th>"
        "<th style='padding:8px 10px;text-align:right;'>Qty</th>"
        "<th style='padding:8px 10px;text-align:center;'>Erect</th>"
        "</tr>"
    )
    return (
        "<div style='max-height:480px;overflow-y:auto;border:1px solid #e3e7ee;border-radius:8px;'>"
        "<table style='width:100%;border-collapse:collapse;font-size:13px;'>"
        f"<thead>{header}</thead><tbody>{''.join(rows_html)}</tbody></table></div>"
    )
 
 
@st.cache_data(max_entries=3)
def read_file(file):
    """Returns (df, error_message). Never raises - a malformed upload
    should show a clean st.error, not crash the whole app."""
    try:
        if file.name.endswith(".csv"):
            df = pd.read_csv(file)
        else:
            df = pd.read_parquet(file)
        df.columns = df.columns.str.strip().str.lower()
        return df, None
    except Exception as e:
        return None, str(e)
 
 
@st.cache_data(show_spinner="Processing data...", max_entries=3)
def process_data(df: pd.DataFrame, cols: dict) -> pd.DataFrame:
    """cols: the resolved {logical_name: real_column_name} mapping picked in the sidebar.
 
    Cached: Streamlit reruns the whole script on every filter/widget change,
    which would otherwise re-run normalize_item() over every row and rebuild
    the category lookup on every single interaction. This only recomputes
    when the uploaded data or the column mapping actually changes."""
    df = df.copy()
    item_col = cols["item_col"]
    qsub_col = cols["qsub_col"]
 
    df["_item_norm"] = df[item_col].apply(normalize_item)
 
    # Raw (un-adjusted) quantity - mirrors what process_cv() reads directly
    # from "qsub" in the export tool, before any HV-multiplier adjustment.
    df["_qsub_raw"] = pd.to_numeric(df[qsub_col], errors="coerce").fillna(0)
 
    # HV pole recovery counts double (applied like the export tool's
    # df.loc[hv_pole_mask, col] *= HV_POLE_MULTIPLIER)
    hv_key_norm = normalize_item(HV_POLE_KEY)
    hv_mask = df["_item_norm"] == hv_key_norm
    df["_qsub_adj"] = df["_qsub_raw"]
    df.loc[hv_mask, "_qsub_adj"] *= HV_POLE_MULTIPLIER
 
    # map every item to its category
    item_to_cat = {}
    for cat_name, mapping in ALL_CATEGORIES.items():
        for desc, label in mapping.items():
            item_to_cat[normalize_item(desc)] = label
    df["_mapped_category"] = df["_item_norm"].map(item_to_cat)
 
    # normalized pole/enid, needed for CV8/CV31 pole-dedupe logic
    pole_col = cols.get("pole_col")
    if pole_col and pole_col in df.columns:
        df["_pole_norm"] = df[pole_col].apply(normalize_pole)
    else:
        df["_pole_norm"] = ""
 
    # cleaned job column
    job_col = cols.get("job_col")
    if job_col and job_col in df.columns:
        df["_job_clean"] = df[job_col].apply(clean_job)
    else:
        df["_job_clean"] = ""
 
    # date column
    date_col = cols.get("date_col")
    df["_date"] = pd.to_datetime(df[date_col], errors="coerce") if date_col and date_col in df.columns else pd.NaT
 
    return df
 
 
@st.cache_data(show_spinner=False, max_entries=5)
def cv7_dedupe_poles(frame: pd.DataFrame) -> set:
    """Poles already covered by a CV7 erect/recover item - mirrors the
    export tool's cv7_poles = cv7_set(df, CV7_erect) | cv7_set(df, CV7_erect_H)
    | cv7_set(df, CV7_erect_lv) | cv7_set(df, CV7_recover)."""
    keys = set()
    for mapping in POLE_CATEGORIES.values():
        keys |= {normalize_item(k) for k in mapping}
    poles = set(frame.loc[frame["_item_norm"].isin(keys), "_pole_norm"].dropna())
    poles.discard("")
    return poles
 
 
def cv_pole_resume(frame: pd.DataFrame, mapping: dict, cv7_poles: set) -> pd.DataFrame:
    """Mirrors the export tool's process_cv(): filter by item, drop
    zero-qty rows, dedupe by pole (keep first), exclude poles already
    counted under CV7. The resulting row count == the exporter's metric."""
    keys = {normalize_item(k) for k in mapping}
    sub = frame[frame["_item_norm"].isin(keys)].copy()
    sub = sub[sub["_qsub_raw"] != 0]
    sub = sub.drop_duplicates(subset="_pole_norm")
    sub = sub[~sub["_pole_norm"].isin(cv7_poles)]
    return sub
 
 
def build_card(frame: pd.DataFrame, cat_name: str, mapping: dict, cv7_poles: set):
    """Returns (cat_name, total_qty, sub) for a single category, filtered
    strictly by that category's OWN item keys - never by the shared
    "_mapped_category" label, since several categories map to the same
    label (e.g. CV7_erect and CV7_erect_H both -> "CV7 HV pole") and
    grouping by label would merge their counts together."""
    if cat_name in POLE_DEDUPE_CATEGORIES:
        sub = cv_pole_resume(frame, mapping, cv7_poles)
        if sub.empty:
            return None
        return (cat_name, len(sub), sub)
    else:
        keys = {normalize_item(k) for k in mapping}
        sub = frame[frame["_item_norm"].isin(keys)]
        if sub.empty:
            return None
        return (cat_name, sub["_qsub_adj"].sum(), sub)
 
 
def build_subtype_card(frame: pd.DataFrame, subtype_name: str, descriptions: list):
    keys = {normalize_item(d) for d in descriptions}
    sub = frame[frame["_item_norm"].isin(keys)]
    if sub.empty:
        return None
    return (subtype_name, sub["_qsub_adj"].sum(), sub)
 
 
def render_metric(slot, cat_name: str, total_qty, display_name: str = None):
    label = display_name or cat_name
    if cat_name in UNIT_CONFIG:
        slot.metric(label, format_length(total_qty, UNIT_CONFIG[cat_name]))
    elif cat_name in POLE_DEDUPE_CATEGORIES:
        slot.metric(label, f"{total_qty:,.0f} poles")
    else:
        slot.metric(label, f"{total_qty:,.0f}")
 
 
# ============================================================
# APP
# ============================================================
st.markdown(
    """
    <style>
    div[data-testid="stMetric"] {
        background-color: #f5f7fa;
        border: 1px solid #e3e7ee;
        border-radius: 10px;
        padding: 14px 16px;
    }
    div[data-testid="stMetricLabel"] { font-weight: 600; }
    </style>
    """,
    unsafe_allow_html=True,
)
 
st.title("⚡ Network Job Tracker Dashboard")
 
uploaded = st.file_uploader("Upload your parquet or CSV file", type=["parquet", "csv"])
if not uploaded:
    st.info("Upload the parquet/CSV file your export script normally reads, then filters and charts appear below.")
    st.stop()
 
raw_df, read_err = read_file(uploaded)
if read_err:
    st.error(f"Couldn't read that file - it may be corrupted or not a valid CSV/parquet file.\n\n**Details:** {read_err}")
    st.stop()
 
with st.expander("Detected columns in your file (click to view)"):
    st.write(list(raw_df.columns))
 
 
def guess(*candidates):
    """Returns the first candidate that exists as a real column, or None if none match."""
    for c in candidates:
        if c in raw_df.columns:
            return c
    return None
 
 
with st.sidebar.expander("⚙️ Column mapping (advanced)", expanded=False):
    st.caption("Confirm each field maps to the right column in your file.")
    col_options = list(raw_df.columns)
    none_option = ["(none)"] + col_options
 
    def pick(label, default_col, key):
        opts = none_option
        idx = opts.index(default_col) if default_col in opts else 0
        if default_col is None:
            st.warning(f"Couldn't guess a column for **{label}** - pick one.")
        val = st.selectbox(label, opts, index=idx, key=key)
        return None if val == "(none)" else val
 
    cols = {
        "item_col": pick("Description / item", guess("item", "description"), "map_item"),
        "qsub_col": pick("Quantity (qsub)", guess("qsub", "quantity_used"), "map_qsub"),
        "district_col": pick("District", guess("shire", "district"), "map_district"),
        "project_col": pick("Project", guess("project"), "map_project"),
        "circuit_col": pick("Circuit", guess("segmentcode", "circuit"), "map_circuit"),
        "pole_col": pick("Pole / enid", guess("pole", "enid"), "map_pole"),
        "pid_col": pick("PID", guess("pid_ohl_nr", "pid"), "map_pid"),
        "total_col": pick("Total value", guess("total"), "map_total"),
        "orig_col": pick("Original value", guess("orig", "original"), "map_orig"),
        "job_col": pick("Job", guess("job", "sourcefile"), "map_job"),
        "date_col": pick("Date", guess("datetouse", "date", "plan1", "done"), "map_date"),
    }
 
missing_required = [k for k in ["item_col", "qsub_col", "district_col", "circuit_col"] if cols[k] is None]
if missing_required:
    st.error(f"These required fields still need a column picked in the sidebar 'Column mapping' section: {missing_required}")
    st.stop()
 
if cols.get("pole_col") is None:
    st.sidebar.warning("No Pole/enid column selected - CV8 and CV31 counts can't be deduplicated by pole and will be shown as raw row counts, which may not match the export tool.")
 
df = process_data(raw_df, cols)
 
district_col, project_col, circuit_col = cols["district_col"], cols["project_col"], cols["circuit_col"]
 
# ---- Sidebar filters ----
st.sidebar.header("🔍 Filters")
 
if df["_date"].notna().any():
    min_d, max_d = df["_date"].min(), df["_date"].max()
    date_range = st.sidebar.date_input("Date", value=(min_d.date(), max_d.date()))
else:
    date_range = None
    st.sidebar.caption("No usable dates found in the selected Date column.")
 
 
def multiselect_filter(label, col, key):
    if not col or col not in df.columns:
        return []
    opts = sorted(df[col].dropna().astype(str).unique())
    return st.sidebar.multiselect(label, opts, key=key)
 
 
districts = multiselect_filter("District", district_col, "f_district")
projects = multiselect_filter("Project", project_col, "f_project")
circuits = multiselect_filter("Circuit", circuit_col, "f_circuit")
 
f = df.copy()
if date_range and isinstance(date_range, tuple) and len(date_range) == 2:
    start, end = pd.Timestamp(date_range[0]), pd.Timestamp(date_range[1]) + pd.Timedelta(days=1) - pd.Timedelta(seconds=1)
    f = f[f["_date"].between(start, end)]
if districts:
    f = f[f[district_col].astype(str).isin(districts)]
if projects:
    f = f[f[project_col].astype(str).isin(projects)]
if circuits:
    f = f[f[circuit_col].astype(str).isin(circuits)]
 
st.sidebar.divider()
st.sidebar.metric("Rows after filters", f"{len(f):,}", delta=f"of {len(df):,} total")
 
tab_overview, tab_jobs, tab_items, tab_forecast, tab_totals = st.tabs(
    ["📊 Overview", "🗂️ Jobs", "📦 Mapped Items", "📈 Pole Position", "💰 Totals"]
)
 
# ---- Overview: CV7_recover over time ----
with tab_overview:
    st.subheader("CV7_recover — count over time")
    recover_keys = {normalize_item(k) for k in CV7_recover}
    recover_df = f[f["_item_norm"].isin(recover_keys)]
    recover_total = recover_df["_qsub_adj"].sum()
    show_total_banner("Total CV7_recover count", f"{recover_total:,.0f}")
 
    if recover_df.empty or recover_df["_date"].isna().all():
        st.caption("No CV7_recover records (with a date) for the current filters.")
    else:
        granularity = st.radio("Group by", ["Day", "Week", "Month"], index=2, horizontal=True)
        freq = {"Day": "D", "Week": "W", "Month": "MS"}[granularity]
        trend = (
            recover_df.dropna(subset=["_date"])
            .set_index("_date")
            .resample(freq)["_qsub_adj"]
            .sum()
            .reset_index()
            .rename(columns={"_date": "Date", "_qsub_adj": "Count"})
        )
        fig = px.bar(trend, x="Date", y="Count", text="Count")
        fig.update_traces(marker_color="#2563eb")
        fig.update_layout(margin=dict(t=10, b=10))
        st.plotly_chart(fig, use_container_width=True)
 
    st.subheader("All pole categories (erect only)")
    # Grouped by SOURCE category (CV7_erect / CV7_erect_H / CV7_erect_lv),
    # not by the shared "_mapped_category" label - several of these
    # categories map to the same label ("CV7 HV pole"), so grouping by
    # label would silently merge their counts. This mirrors how the
    # export tool's Summary sheet keeps each category as its own column.
    # CV7_recover is excluded here since it already has its own chart above.
    pole_rows = []
    for cat_name, mapping in POLE_CATEGORIES.items():
        if cat_name == "CV7_recover":
            continue
        keys = {normalize_item(k) for k in mapping}
        sub = f[f["_item_norm"].isin(keys)]
        if not sub.empty:
            pole_rows.append({
                "Pole type": POLE_DISPLAY_NAMES.get(cat_name, cat_name),
                "Count": sub["_qsub_adj"].sum(),
            })
    pole_summary = pd.DataFrame(pole_rows)
 
    pole_total = pole_summary["Count"].sum() if not pole_summary.empty else 0
    show_total_banner("Total poles (all categories)", f"{pole_total:,.0f}")
 
    if not pole_summary.empty:
        fig2 = px.bar(pole_summary, x="Pole type", y="Count", text="Count", color="Pole type")
        fig2.update_layout(showlegend=False, margin=dict(t=10, b=10))
        st.plotly_chart(fig2, use_container_width=True)
    else:
        st.caption("No pole records for the current filters.")
 
# ---- Jobs tab ----
with tab_jobs:
    # ---- Outages Programme (uploaded workbook, cached on file content) ----
    st.subheader("Outages Programme 2026")
 
    outage_upload = st.file_uploader(
        "Upload High-level_planning_2026.xlsx",
        type=["xlsx"],
        key="outage_file_uploader",
    )
 
    if outage_upload is None:
        st.info("Upload the outages programme workbook (sheet '2026', header on row 7) to see it here.")
        outage_df = None
    else:
        outage_df, outage_err = load_outage_programme(outage_upload.getvalue())
        if outage_err:
            st.error(
                f"Couldn't read that workbook - check it has a sheet named '2026' with headers on row 7.\n\n"
                f"**Details:** {outage_err}"
            )
            outage_df = None
 
    if outage_df is not None:
        st.caption(f"{len(outage_df):,} rows from High-level_planning_2026.xlsx (sheet '2026')")
 
        oc1, oc2, oc3 = st.columns(3)
        with oc1:
            outage_districts = st.multiselect(
                "District", sorted(outage_df["District"].dropna().unique()), key="outage_district"
            )
        with oc2:
            outage_pms = st.multiselect(
                "SPEN PM", sorted(outage_df["SPEN PM"].dropna().unique()), key="outage_pm"
            )
        with oc3:
            if outage_df["Outage Date"].notna().any():
                min_od, max_od = outage_df["Outage Date"].min(), outage_df["Outage Date"].max()
                outage_date_range = st.date_input(
                    "Outage date", value=(min_od.date(), max_od.date()), key="outage_date_range"
                )
            else:
                outage_date_range = None
 
        outage_f = outage_df.copy()
        if outage_districts:
            outage_f = outage_f[outage_f["District"].isin(outage_districts)]
        if outage_pms:
            outage_f = outage_f[outage_f["SPEN PM"].isin(outage_pms)]
        if outage_date_range and isinstance(outage_date_range, tuple) and len(outage_date_range) == 2:
            o_start = pd.Timestamp(outage_date_range[0])
            o_end = pd.Timestamp(outage_date_range[1]) + pd.Timedelta(days=1) - pd.Timedelta(seconds=1)
            outage_f = outage_f[outage_f["Outage Date"].between(o_start, o_end)]
 
        st.caption(f"{len(outage_f):,} rows after outage filters")
 
        dl_col1, dl_col2, dl_col3 = st.columns([1.4, 1.1, 2])
        with dl_col1:
            st.download_button(
                "📥 Full calendar (Excel, calendar layout)",
                data=build_full_calendar_excel(outage_f),
                file_name=f"outages_calendar_{datetime.now():%Y%m%d}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                help="One sheet per month laid out like the calendar above, every outage's full name written out in full - plus a flat 'All outages' list sheet.",
            )
        with dl_col2:
            csv_export = outage_f.sort_values("Outage Date").assign(
                **{"Full name": outage_f.apply(_outage_full_title, axis=1)}
            )
            st.download_button(
                "📥 Full list (CSV)",
                data=csv_export.to_csv(index=False).encode("utf-8"),
                file_name=f"outages_list_{datetime.now():%Y%m%d}.csv",
                mime="text/csv",
            )
        with dl_col3:
            st.caption("Both downloads use the outage filters above (District / SPEN PM / date) and spell out every event's full name in full - nothing clipped or hidden behind a hover.")
 
        outage_view = st.radio("View", ["Table", "Calendar"], index=1, horizontal=True, key="outage_view")
 
        if outage_view == "Table":
            st.dataframe(outage_f, height=420, use_container_width=True, hide_index=True)
        else:
            cal_df = outage_f.dropna(subset=["Outage Date"])
            events = build_calendar_events(cal_df)
 
            calendar_options = {
                "initialView": "dayGridMonth",
                "headerToolbar": {
                    "left": "prev,next today",
                    "center": "title",
                    "right": "dayGridMonth,listMonth",
                },
                # "auto" instead of a fixed pixel height: dayMaxEvents:False (below)
                # lets every day cell grow tall enough to show ALL of its events
                # directly, so nothing is hidden behind a "+N more" link that would
                # otherwise need a click to open. A fixed height would just clip or
                # squash that content instead of revealing it.
                "height": "auto",
                "firstDay": 1,
                # False = no per-day cap - every outage for every day renders inline,
                # so information is visible with no interaction at all (this trades
                # away the old "+N more" collapsing, which kept row heights uniform
                # and rendering cheap on very busy days - worth knowing if a month
                # with dozens of outages on one day starts to feel sluggish).
                "dayMaxEvents": False,
            }
 
            # Keeps each event on a single line (ellipsis instead of wrapping onto a
            # second/third line and breaking the day cell's height/layout). The full
            # text is still there in the "title" attribute above, so hovering an
            # event reveals it regardless of how it's visually clipped here. The
            # streamlit-calendar wrapper only accepts plain JSON options (no custom
            # JS), so a true "expand on hover" handler isn't available - showing
            # every event inline (dayMaxEvents:False above) is the closest
            # equivalent: nothing is hidden behind an interaction in the first place.
            calendar_custom_css = """
                .fc-event-title { white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }
                .fc-daygrid-event { white-space: nowrap; cursor: pointer; }
                .fc-daygrid-day-number { cursor: pointer; }
                .fc-daygrid-day:hover { background-color: #eef2ff; }
            """
 
            # dateClick/eventClick re-enabled (each click now costs a Streamlit
            # rerun) so a click can drive the breakdown + Work Instructions table
            # below - callbacks=[] previously traded that away for speed. The
            # heavier per-rerun computations elsewhere (cv7_dedupe_poles, this
            # event list, the pole/task table) are now cached, so a click here
            # no longer re-runs them from scratch every time.
            calendar_result = st_calendar(
                events=events,
                options=calendar_options,
                custom_css=calendar_custom_css,
                callbacks=["dateClick", "eventClick"],
                key="outage_calendar",
            )
 
            clicked_date = None
            if calendar_result.get("eventClick"):
                ev_start = calendar_result["eventClick"].get("event", {}).get("start")
                if ev_start:
                    clicked_date = pd.to_datetime(ev_start).date()
            elif calendar_result.get("dateClick"):
                dc = calendar_result["dateClick"]
                dc_date = dc.get("date") or dc.get("dateStr")
                if dc_date:
                    clicked_date = pd.to_datetime(dc_date).date()
 
            if clicked_date is not None:
                st.session_state["selected_outage_date"] = clicked_date
 
            selected_date = st.session_state.get("selected_outage_date")
 
            st.divider()
            if selected_date is None:
                st.caption("Click a date or an outage above to see its full breakdown and Work Instructions here.")
            else:
                day_rows = outage_f[outage_f["Outage Date"].dt.date == selected_date]
                st.markdown(f"**Outages on {selected_date:%d %b %Y}** ({len(day_rows):,})")
                if day_rows.empty:
                    st.caption("No outages on this date under the current filters.")
                else:
                    st.dataframe(day_rows, use_container_width=True, hide_index=True)
 
                st.markdown(f"**Work Instructions — poles & MD Poling tasks on {selected_date:%d %b %Y}**")
                if cols.get("date_col") is None or f["_date"].isna().all():
                    st.caption(
                        "No usable dates in the main dataset's mapped 'Date' column - pick one under "
                        "'Column mapping' in the sidebar to enable this breakdown."
                    )
                else:
                    day_job_rows = f[f["_date"].dt.date == selected_date]
                    pole_task_df = build_pole_task_table(
                        day_job_rows, cols["item_col"], cols["qsub_col"], cols.get("pole_col")
                    )
                    if pole_task_df.empty:
                        st.caption("No pole/task records for this date in the main dataset under the current filters.")
                    else:
                        st.caption(f"{pole_task_df['Pole'].nunique():,} pole(s) · {len(pole_task_df):,} task row(s)")
                        st.markdown(render_pole_task_table_html(pole_task_df), unsafe_allow_html=True)
 
# ---- Mapped items tab: image-led groups, then the rest as a card grid ----
with tab_items:
    st.subheader("Mapped items")
 
    cv7_poles = cv7_dedupe_poles(f)
    all_card_data = []  # accumulated in display order, feeds the detail selectbox below
 
    grouped_cat_names = {c for group in CARD_GROUPS for c in group.get("categories", [])}
    grouped_cat_names |= {c for group in CARD_GROUPS if "subtypes" in group for c in ["Switch"]}
 
    for group in CARD_GROUPS:
        title = group["title"]
        img_l, img_c, img_r = st.columns([1, 1, 1])
        with img_c:
            st.markdown(f"<h3 style='text-align:center; margin-bottom:0.3rem;'>{title}</h3>", unsafe_allow_html=True)
            image_path = group.get("image")
            if image_path and os.path.exists(image_path):
                st.image(image_path, width=300)
            elif image_path:
                st.caption(f"⚠️ Image not found: {image_path}")
 
        group_cards = []
 
        if "subtypes" in group:
            for subtype_name, descriptions in group["subtypes"].items():
                card = build_subtype_card(f, subtype_name, descriptions)
                if card:
                    group_cards.append(card)
                    all_card_data.append(card)
        else:
            for cat_name in group.get("categories", []):
                mapping = ALL_CATEGORIES.get(cat_name)
                if mapping is None:
                    continue
                card = build_card(f, cat_name, mapping, cv7_poles)
                if card:
                    group_cards.append(card)
                    all_card_data.append(card)
 
        if group_cards:
            row_cols = st.columns(len(group_cards))
            for slot, (cat_name, total_qty, _sub) in zip(row_cols, group_cards):
                render_metric(slot, cat_name, total_qty, display_name=POLE_DISPLAY_NAMES.get(cat_name))
        else:
            st.caption("No records for this group under the current filters.")
 
        st.divider()
 
    # Everything not already shown in an image group, same grid as before
    remaining_cat_names = [c for c in ALL_CATEGORIES if c not in grouped_cat_names]
    remaining_cards = []
    for cat_name in remaining_cat_names:
        card = build_card(f, cat_name, ALL_CATEGORIES[cat_name], cv7_poles)
        if card:
            remaining_cards.append(card)
            all_card_data.append(card)
 
    if remaining_cards:
        st.markdown("**Other items**")
        n_cols = 4
        rows = [remaining_cards[i:i + n_cols] for i in range(0, len(remaining_cards), n_cols)]
        for row in rows:
            row_cols = st.columns(n_cols)
            for slot, (cat_name, total_qty, _sub) in zip(row_cols, row):
                render_metric(slot, cat_name, total_qty)
 
    if not all_card_data:
        st.caption("No mapped items for the current filters.")
    else:
        st.divider()
        chosen = st.selectbox("View details for", [c[0] for c in all_card_data])
        _, _, sub = next(c for c in all_card_data if c[0] == chosen)
        detail = pd.DataFrame({
            "District": sub[district_col],
            "Job": sub["_job_clean"],
            "Circuit": sub[circuit_col],
            "enid": sub[cols["pole_col"]] if cols["pole_col"] in sub.columns else "",
        })
        st.dataframe(detail, height=320, use_container_width=True, hide_index=True)
 
# ---- Poles Forecast tab ----
@st.cache_data(show_spinner="Reading poles forecast workbook...", max_entries=3)
def list_forecast_sheets(file_bytes: bytes):
    """Returns (sheet_names, error_message)."""
    try:
        return pd.ExcelFile(io.BytesIO(file_bytes)).sheet_names, None
    except Exception as e:
        return [], str(e)
 
 
@st.cache_data(show_spinner="Reading poles forecast workbook...", max_entries=3)
def load_forecast_workbook(file_bytes: bytes, sheet_name: str):
    """Returns (df, error_message)."""
    try:
        fdf = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name)
        fdf.columns = fdf.columns.astype(str).str.strip()
        return fdf, None
    except Exception as e:
        return None, str(e)
 
 
with tab_forecast:
    st.subheader("Pole Position")
 
    forecast_upload = st.file_uploader(
        "Upload poles forecast workbook (.xlsx)",
        type=["xlsx"],
        key="forecast_file_uploader",
    )
 
    if forecast_upload is None:
        st.info(
            "Upload the workbook with columns: District, Project ID, Project, "
            "Circuit, Voltage, Forecasted Total poles, Poles Disposed."
        )
    else:
        f_bytes = forecast_upload.getvalue()
        sheet_names, sheet_err = list_forecast_sheets(f_bytes)
        if sheet_err:
            st.error(f"Couldn't open that workbook.\n\n**Details:** {sheet_err}")
            sheet_names = []
 
        if not sheet_names:
            pass  # error already shown above - nothing more to render in this tab
        else:
            sheet_choice = (
                st.selectbox("Sheet", sheet_names, key="forecast_sheet")
                if len(sheet_names) > 1 else sheet_names[0]
            )
            fdf, fdf_err = load_forecast_workbook(f_bytes, sheet_choice)
            if fdf_err:
                st.error(f"Couldn't read sheet '{sheet_choice}'.\n\n**Details:** {fdf_err}")
                fdf = None
 
            if fdf is None:
                pass  # error already shown above
            else:
                with st.expander("Detected columns (click to view)"):
                    st.write(list(fdf.columns))
 
                def fguess(*candidates):
                    lower_map = {c.lower(): c for c in fdf.columns}
                    for cand in candidates:
                        if cand.lower() in lower_map:
                            return lower_map[cand.lower()]
                    return None
 
                fcol_options = ["(none)"] + list(fdf.columns)
 
                with st.expander("⚙️ Column mapping", expanded=False):
                    def fpick(label, default_col, key):
                        idx = fcol_options.index(default_col) if default_col in fcol_options else 0
                        if default_col is None:
                            st.warning(f"Couldn't guess a column for **{label}** - pick one.")
                        val = st.selectbox(label, fcol_options, index=idx, key=key)
                        return None if val == "(none)" else val
 
                    f_cols = {
                        "district": fpick("District", fguess("District"), "fc_district"),
                        "pid": fpick("Project ID", fguess("Project ID", "PID"), "fc_pid"),
                        "project": fpick("Project", fguess("Project", "Project Name"), "fc_project"),
                        "circuit": fpick("Circuit", fguess("Circuit"), "fc_circuit"),
                        "voltage": fpick("Voltage", fguess("Voltage"), "fc_voltage"),
                        "forecast": fpick(
                            "Forecasted Total poles",
                            fguess("Forecasted Total poles", "Forecasted Total Poles", "Forecast Total Poles"),
                            "fc_forecast",
                        ),
                        "disposed": fpick("Poles Disposed", fguess("Poles Disposed", "Poles disposed"), "fc_disposed"),
                        "start_date": fpick("Start Date", fguess("Start Date", "StartDate", "Start"), "fc_start_date"),
                    }
 
                required = ["project", "circuit", "pid", "forecast", "disposed"]
                missing = [k for k in required if f_cols[k] is None]
                if missing:
                    st.error(f"Please map these columns in 'Column mapping' above: {missing}")
                else:
                    plot_df = pd.DataFrame({
                        "District": fdf[f_cols["district"]] if f_cols["district"] else "",
                        "PID": fdf[f_cols["pid"]],
                        "Project": fdf[f_cols["project"]],
                        "Circuit": fdf[f_cols["circuit"]],
                        "Voltage": fdf[f_cols["voltage"]] if f_cols["voltage"] else "",
                        "Forecast": pd.to_numeric(fdf[f_cols["forecast"]], errors="coerce").fillna(0),
                        "Disposed": pd.to_numeric(fdf[f_cols["disposed"]], errors="coerce").fillna(0),
                    })
                    if f_cols["start_date"]:
                        plot_df["Start Date"] = pd.to_datetime(fdf[f_cols["start_date"]], errors="coerce")
                        plot_df["Year"] = plot_df["Start Date"].dt.year
                    plot_df = plot_df.dropna(subset=["Project"])
                    # clamp so a data-entry error (disposed > forecasted) never draws past the bar
                    plot_df["Disposed"] = plot_df[["Disposed", "Forecast"]].min(axis=1)
                    plot_df["Remaining"] = (plot_df["Forecast"] - plot_df["Disposed"]).clip(lower=0)
                    plot_df["Label"] = (
                        plot_df["Project"].astype(str) + " — "
                        + plot_df["Circuit"].astype(str) + " — PID "
                        + plot_df["PID"].astype(str)
                    )
 
                    fc1, fc2, fc3 = st.columns(3)
                    with fc1:
                        forecast_districts = (
                            st.multiselect("District", sorted(plot_df["District"].dropna().unique()), key="forecast_district")
                            if f_cols["district"] else []
                        )
                    with fc2:
                        forecast_voltages = (
                            st.multiselect("Voltage", sorted(plot_df["Voltage"].dropna().unique()), key="forecast_voltage")
                            if f_cols["voltage"] else []
                        )
                    with fc3:
                        forecast_years = (
                            st.multiselect(
                                "Start year",
                                sorted(plot_df["Year"].dropna().unique().astype(int)),
                                key="forecast_year",
                            )
                            if "Year" in plot_df.columns else []
                        )
 
                    if forecast_districts:
                        plot_df = plot_df[plot_df["District"].isin(forecast_districts)]
                    if forecast_voltages:
                        plot_df = plot_df[plot_df["Voltage"].isin(forecast_voltages)]
                    if forecast_years:
                        plot_df = plot_df[plot_df["Year"].isin(forecast_years)]
 
                    total_forecast = plot_df["Forecast"].sum()
                    total_disposed = plot_df["Disposed"].sum()
                    show_total_banner(
                        "Poles disposed vs forecasted",
                        f"{total_disposed:,.0f} / {total_forecast:,.0f}"
                        + (f"  ({total_disposed / total_forecast:.0%})" if total_forecast else ""),
                    )
 
                    if plot_df.empty:
                        st.caption("No rows to chart for the current filters.")
                    else:
                        plot_df = plot_df.sort_values("Forecast", ascending=True)
 
                        fig = go.Figure()
                        fig.add_trace(go.Bar(
                            y=plot_df["Label"], x=plot_df["Disposed"], orientation="h",
                            name="Disposed", marker_color="#16a34a",
                            hovertemplate="%{y}<br>Disposed: %{x:,.0f}<extra></extra>",
                        ))
                        fig.add_trace(go.Bar(
                            y=plot_df["Label"], x=plot_df["Remaining"], orientation="h",
                            name="Remaining", marker_color="#dc2626",
                            hovertemplate="%{y}<br>Remaining: %{x:,.0f}<extra></extra>",
                        ))
                        # Total (= Forecast = Disposed + Remaining) printed just past the end
                        # of each stacked bar, so the project total is readable at a glance
                        # without having to add the two segments up by eye.
                        fig.add_trace(go.Scatter(
                            y=plot_df["Label"], x=plot_df["Forecast"],
                            mode="text",
                            text=[f"{v:,.0f}" for v in plot_df["Forecast"]],
                            textposition="middle right",
                            textfont=dict(size=12, color="#1e293b"),
                            showlegend=False,
                            hoverinfo="skip",
                        ))
                        max_forecast = plot_df["Forecast"].max()
                        fig.update_layout(
                            barmode="stack",
                            height=max(420, 34 * len(plot_df)),
                            margin=dict(l=10, r=60, t=10, b=10),
                            legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
                            xaxis=dict(title="Poles", range=[0, max_forecast * 1.15 if max_forecast else 1]),
                        )
                        st.plotly_chart(fig, use_container_width=True)
                        st.caption(f"{len(plot_df):,} project/circuit rows shown")
 
 
with tab_totals:
    GBP_COLUMN_CONFIG = lambda col_name: {col_name: st.column_config.NumberColumn(col_name, format="£%.2f")}
 
    total_val = pd.to_numeric(f[cols["total_col"]], errors="coerce").sum() if cols["total_col"] in f.columns else None
    orig_val = pd.to_numeric(f[cols["orig_col"]], errors="coerce").sum() if cols["orig_col"] in f.columns else None
 
    if total_val is not None:
        show_total_banner("Total value (£)", f"£{total_val:,.2f}")
 
    c1, c2 = st.columns(2)
    if total_val is not None:
        c1.metric("Total value", f"£{total_val:,.2f}")
    if total_val is not None and orig_val is not None:
        c2.metric("Difference vs original", f"£{total_val - orig_val:,.2f}")
 
    if total_val is not None and project_col in f.columns:
        st.divider()
        st.subheader("Total value by Project")
        f["_total_val_row"] = pd.to_numeric(f[cols["total_col"]], errors="coerce")
        by_project_total = (
            f.groupby(project_col)["_total_val_row"].sum()
            .reset_index()
            .rename(columns={project_col: "Project", "_total_val_row": "Total value (£)"})
            .sort_values("Total value (£)", ascending=False)
        )
        st.caption(f"{len(by_project_total):,} projects under the current filters")
 
        fig_proj = px.bar(
            by_project_total.sort_values("Total value (£)", ascending=True),
            x="Total value (£)", y="Project", orientation="h", text="Total value (£)",
        )
        fig_proj.update_traces(marker_color="#2563eb", texttemplate="£%{text:,.0f}")
        fig_proj.update_layout(height=max(360, 32 * len(by_project_total)), margin=dict(l=10, r=10, t=10, b=10))
        st.plotly_chart(fig_proj, use_container_width=True)
 
        st.dataframe(
            by_project_total, height=320, use_container_width=True, hide_index=True,
            column_config=GBP_COLUMN_CONFIG("Total value (£)"),
        )
 
    if total_val is not None and orig_val is not None:
        f["_row_variance"] = pd.to_numeric(f[cols["total_col"]], errors="coerce") - pd.to_numeric(f[cols["orig_col"]], errors="coerce")
        variance_rows = f[f["_row_variance"] != 0]
 
        st.subheader("Jobs where total ≠ original")
        variance_table = pd.DataFrame({
            "District": variance_rows[district_col],
            "Job": variance_rows["_job_clean"],
            "Circuit": variance_rows[circuit_col],
            "Difference (£)": variance_rows["_row_variance"],
        })
        st.caption(f"{len(variance_table):,} rows with a variance under the current filters")
        st.dataframe(
            variance_table.sort_values("Difference (£)", key=abs, ascending=False),
            height=320, use_container_width=True, hide_index=True,
            column_config=GBP_COLUMN_CONFIG("Difference (£)"),
        )
 
        st.subheader("Difference by Job")
        by_job = (
            variance_rows.assign(Job=variance_rows["_job_clean"])
            .groupby("Job")["_row_variance"].sum()
            .reset_index()
            .rename(columns={"_row_variance": "Difference (£)"})
            .sort_values("Difference (£)", key=abs, ascending=False)
        )
        st.dataframe(
            by_job, height=280, use_container_width=True, hide_index=True,
            column_config=GBP_COLUMN_CONFIG("Difference (£)"),
        )
 
        st.subheader("Difference by Project")
        if project_col in variance_rows.columns:
            by_project = (
                variance_rows.groupby(project_col)["_row_variance"].sum()
                .reset_index()
                .rename(columns={project_col: "Project", "_row_variance": "Difference (£)"})
                .sort_values("Difference (£)", key=abs, ascending=False)
            )
            st.dataframe(
                by_project, height=280, use_container_width=True, hide_index=True,
                column_config=GBP_COLUMN_CONFIG("Difference (£)"),
            )
